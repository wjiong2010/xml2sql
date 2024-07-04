"""
File Name: excel_format.py
Author: John.Wang
Date: 20240604
Description: Excel cell format
"""
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

# from openpyxl.utils import get_column_letter, column_index_from_string


"""
    白色：FFFFFF，黑色：000000，红色：FF0000，黄色：FFFF00
    绿色：00FF00，蓝色：0000FF，橙色：FF9900，灰色：C0C0C0
    常见颜色代码表：https://www.osgeo.cn/openpyxl/styles.html#indexed-colours
"""
COLOR_WHITE = 'FFFFFF'
COLOR_BLACK = '000000'
COLOR_RED = 'FF0000'
COLOR_GREEN = '00FF00'
COLOR_BLUE = '0000FF'
COLOR_YELLOW = 'FFFF00'
COLOR_GRAY = 'C0C0C0'
COLOR_ORANGE = 'FF9900'


class ExcelFormat:
    def __init__(self):
        """设置单元格文字样式"""
        self.cell_content_font = Font(bold=False,  # 加粗
                                      italic=False,  # 倾斜
                                      name="Calibri",  # 字体
                                      size=13,  # 文字大小
                                      color=COLOR_BLACK,  # 字体颜色为黑
                                      # underline='single'  # 下划线
                                      )

        self.cell_title_font = Font(bold=True,  # 加粗
                                    italic=False,  # 倾斜
                                    name="Calibri",  # 字体
                                    size=13,  # 文字大小
                                    color=COLOR_BLACK,  # 字体颜色为黑
                                    # underline='single'  # 下划线
                                    )

        """设置单元格边框为黑色边框"""
        self.cell_border = Border(bottom=Side(style='thin', color='000000'),
                                  right=Side(style='thin', color='000000'),
                                  left=Side(style='thin', color='000000'),
                                  top=Side(style='thin', color='000000'))

        """设置单元格对齐方式为水平居中和垂直居中，单元格内容超出范围自动换行"""
        self.cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        """设置单元格底纹颜色"""
        self.cell_fill = PatternFill(fill_type='solid', start_color=COLOR_WHITE)

        self.row_height = 20
        self.column_height = 15

    """row"""

    def set_row(self, ws, row, height=0):
        row = ws.row_dimensions[row]
        # 设置行高
        if row == 0:
            row.height = self.row_height
        else:
            row.height = height

    """column"""

    def set_column(self, ws, column, width=0):
        column = ws.column_dimensions[column]
        # column = ws.column_dimensions[get_column_letter(1)]  # 根据数字列标获取第一列列对象
        if width == 0:
            column.width = self.column_height  # 设置列宽
        else:
            column.width = width

    def set_cell(self, cell, font=None, align=None, border=None):
        if font is None:
            cell.font = self.cell_content_font
        else:
            cell.font = font
        cell.border = self.cell_border
        cell.alignment = self.cell_alignment
        cell.fill = self.cell_fill
